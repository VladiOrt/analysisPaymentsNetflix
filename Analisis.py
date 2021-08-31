import calendar
from re import U
from numpy.lib.function_base import percentile
from numpy.lib.shape_base import split
import openpyxl
from numpy import array
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook




#Guardamos en variables nuestros documentos
wbNetflix = load_workbook('NET.xlsx', data_only=True) #Documento del archivo a comparar
BD = load_workbook('Audiomaster.xlsx') #Archivo de Tareas realizadas en la empresa
Task  = load_workbook('Task.xlsx') #Base de datos de Tipo de tareas que se pueden realizar

wSNetflix = wbNetflix.worksheets[0] #Primer Hoja del archivo a comparar
wSBD = BD.worksheets[0] #Primer Hoja de la BD de tareas realizadas
wSTask = Task.worksheets[0]  #BD de Tipo de Tareas

numeroFila=15 #Variable para el numero de filas en las que nos encontramos 
UltimaFIla=0
FechaInicial = ''
FechaMenor = ''
FechaMayor = ''

#Primer Ciclo para recorrer el documento en excel
for row in enumerate(wSNetflix.rows): #Para cada celda en el archivo excel    
    ########### CODIGO PARA TERMINAR EL ANALISIS EN CASO DE QUE YA NO HAYA INFORMACION ########
    #Obtenemos la primer celda para verificar si hay que analizar
    valorPrimerCelda = wSNetflix.cell(numeroFila,1).value
    valorPrimerCelda = str(valorPrimerCelda)
    valorPrimerCelda = valorPrimerCelda.replace(" ","")    #Adecuamos la informacion de la celda
    
    if(valorPrimerCelda == "RemittanceTotal:") : #Si el valor de la celda es igual al total
        UltimaFIla = numeroFila        
        break  #Terminamos el proceso    
    ########## FIN DE CONDIGO PARA TERMINAR ANALISIS  ##########################################
    ######### DETERMINAR FECHA MENOR Y MAYOR  ##############################
    if wSNetflix.cell(numeroFila,19).value != "":
        if numeroFila ==15 :
            valorFecha = wSNetflix.cell(numeroFila,3).value
            valorFecha = valorFecha.split("||")
            valorFecha = valorFecha[(len(valorFecha)-1)]
            FechaMenor = valorFecha
            FechaMayor = valorFecha
        if numeroFila >15: 
            valorFecha = wSNetflix.cell(numeroFila,3).value
            valorFecha = str(valorFecha)
            valorFecha = valorFecha.split("||")
            valorFecha = valorFecha[(len(valorFecha)-1)]
            FechaMenorActual = valorFecha
            FechaMayorActual = valorFecha
            if FechaMenorActual < FechaMenor:
                FechaMenor = FechaMenorActual
            if FechaMayorActual > FechaMayor:
                FechaMayor = FechaMayorActual
    ##########   FIN DE FECHA MENOR Y MAYOR   ################################
    ################ INICIO DE ANALISIS DE ARCHIVO  ############################################    
    #Separamos la Descripcion 
    if numeroFila ==15 :
        FechaInicial = wSNetflix.cell(numeroFila,2).value 
    Texto = wSNetflix.cell(numeroFila,3).value
    Texto = str(Texto)    
    Valores = Texto.split('||')
    Coincidencias = 0
    ID_uno = ''
    Date_uno = ''
    Date_Init_uno = ''
    Date_End_uno = ''
    Longitud = len(Valores)
    
    #Guardamos las Descripcion individial en las celdas.
    for i in range(1,(Longitud+1)):
        wSNetflix.cell(numeroFila,5+i, Valores[i-1])

    #Guardamos el Movie ID en una variable y la adaptamos para trabajar con ella                
    ID_uno = wSNetflix.cell(numeroFila,10).value
    Date_uno = wSNetflix.cell(numeroFila,19).value
    ID_uno = (str(ID_uno)).strip()
    ID_uno = ID_uno.split(' ')
    ID_uno = ID_uno[(len(ID_uno)-1)]
    ID_uno = ID_uno.strip()

    #Obtenemos la primer fecha y posterior  Sumamos dos dia y restamos dos dias para el intervalo de fechas
    Date_uno = Date_uno.strip()
    Date_uno = datetime.strptime(Date_uno, '%Y-%m-%d')
    Date_Init_uno = Date_uno - timedelta(days=2)
    Date_Init_uno = str(Date_Init_uno)
    Date_Init_uno = Date_Init_uno.strip()
    Date_End_uno = Date_uno + timedelta(days=3)
    Date_End_uno = str(Date_End_uno)
    Date_End_uno = Date_End_uno.strip()

    #Variable para recorrer la BD de tareas.
    filaBDTareas = 1
    ID_uno = ID_uno.strip()
    ID_uno = str(ID_uno)
    
    CoincidenciasId = [None] * 10 #Creamos una lista con valores vacios en caso de que haya mas coincidencias

    #Recorremos la BD de tareas internas
    for row in enumerate(wSBD.rows):
        ID_dos = wSBD.cell(filaBDTareas,25).value
        Date_dos = wSBD.cell(filaBDTareas,19).value
        ID_dos = str(ID_dos)
        EsHTTP = ID_dos.find("requestID")
        if EsHTTP != -1:
            ID_dos = ID_dos.split('requestID=')
            ID_dos = ID_dos[1]
            Date_dos = str(Date_dos)
            ID_dos = ID_dos.strip()

            if (ID_uno == ID_dos) and (Date_dos == Date_uno): 
                CoincidenciasId[Coincidencias] = numeroFila,filaBDTareas
                Coincidencias = Coincidencias+1            
                break
            if ID_uno == ID_dos and (Date_dos >= Date_Init_uno  and Date_dos <= Date_End_uno): 
                CoincidenciasId[Coincidencias] = numeroFila,filaBDTareas
                Coincidencias = Coincidencias+1        
        filaBDTareas = filaBDTareas + 1
        #  COMPARACION CON LAS FILAS COINCIDENTES #
    corrimientoComparacion = 1
    Coincide = 0
    for index in CoincidenciasId: #CORRIMIENTO DE ARRAY
        
        if ((index != None) and (Coincide == 0)): #SI NO ES VACIA LA COINCIDENCIA                                                 
            # INICIO PARA EL CALCULO DE DIFERENCIA DE TIEMPO
            FilaNet = index[0]
            FilaAud = index[1]
            #ValorAsset = wSNetflix.cell(FilaNet,16).value        
            Duracion1 = wSNetflix.cell(FilaNet,16).value
            Duracion1 = Duracion1.strip()
            Duracion1 = Duracion1.split(' ')
            Duracion1 = int(Duracion1[0])   
            Duracion2 = wSBD.cell(FilaAud,8).value   
            if Duracion2==None:                
                Duracion2 = wSBD.cell(FilaAud,7).value   
            Duracion2 = int(Duracion2)
        
            DiferenciaTiempo = Duracion1 - Duracion2
            Fecha1 = wSBD.cell(FilaAud,19).value
            Fecha2 = wSNetflix.cell(FilaNet,19).value 
            Fecha2 = Fecha2 + " 00:00:00"
           
          
            #wSNetflix.cell(FilaNet,corrimiento_Comparacion + 20, (DiferenciaTiempo))
            # FIN DE CALCULO PARA DIFERENCIA DE TIEMPO 

            Fecha = wSBD.cell(FilaAud,19).value
            NombreAsset = wSBD.cell(FilaAud,1).value
            #Desglozamos las intervenciones que deberian aparecer
            Tareas = wSBD.cell(FilaAud,16).value
            Tareas = Tareas.split('/')
            NumTareas =str(len(Tareas))

            Busqueda = 15
            NumeroTareas = 0
            Faltantes = ""
            Match = 0  #

            # Recorremos ahora el archivo de tareas de Netflix para buscar el numero de tareas que hay 
            for row in enumerate(wSNetflix.rows):
                ####### FINALIZAMOS EL FOR EN CASO DE QUE LLEGUEMOS A LA ULTIMA FILA 
                valorPrimerCelda = wSNetflix.cell(Busqueda,1).value
                valorPrimerCelda = valorPrimerCelda.replace(" ","")    #Adecuamos la informacion de la celda
                if(valorPrimerCelda == "RemittanceTotal:" ): #En caso de que lleguemos a la ultima fila                 
                    break # Terminamos el ciclo for
                
                ####### FIN DE CONDICION PARA FINALIZAR
                
                #Obtenemos el ID dentro de Dashboard
                ID = wSNetflix.cell(Busqueda,3).value                      
                ID = ID.split('||')          
                DesIndividual = ID[8].strip()                                  
                ID = ID[4].strip()                    
                ID = ID.replace('Request ID -',"")      
                ID = ID.strip()
                DesTareas = Tareas
                #Contamos cuantas veces se repite el ID en el A.Netflix             
                if ID_uno == ID:
                    
                    NumeroTareas = NumeroTareas +1    
                    DesNet = '' #Descripcion de la tarea por parte de Netflix
                    Opcion2 = ''
                    BuscarTarea = 1
                    for row in enumerate(wSTask.rows): #Buscamos las tareas en la BD de tareas
                        if (wSTask.cell(BuscarTarea,1).value == DesIndividual ):#Si las descripciones son =
                            DesNet = wSTask.cell(BuscarTarea,4).value   #Obtenemos las opciones con las
                            Opcion2 = wSTask.cell(BuscarTarea,6).value   #que podemos compararlas
                            break
                        BuscarTarea = BuscarTarea + 1    
                    #Recorremos el numero de tareas para ver si coinciden para dejar solo las que faltan                  
                    for i in range(0,int(NumTareas)):                                                           
                            if ((DesTareas[i] == DesNet) or (DesTareas[i] == Opcion2)):
                                DesTareas[i] = None 

                Busqueda = Busqueda + 1
                
            CompararTarea = ''


            
    
            #Si el numero de tareas tenemos que debemos porner OK
            #Obtenemos la diferencia en el numero de tareas
            if NumeroTareas == NumTareas:
                CompararTarea = "Ok"
            else: 
                Total = int(NumTareas) - int(NumeroTareas)
                Total = str(Total)                
                DesTareas = str(DesTareas)
                if Total == "0":
                    CompararTarea = "Numero de tareas: OK "
                else:
                    CompararTarea = "Faltan " + Total + " Tareas "+ DesTareas 
            #COMPARAMOS LAS FECHAS PARA DESCARTAR       
            compararFecha = 0
            if str(Fecha1) == str(Fecha2):
                compararFecha = 1
            
            #ANALIZAMOD SI ES RUSH O NO            
            Rush = ""
            valorAsset = wSBD.cell(FilaAud,1).value
            valorAsset = str(valorAsset)
            valorRush = wSBD.cell(FilaAud,15).value
            valorRush = valorRush.strip()
            if (valorRush == "High/Rush"):
                Rush = "SI"
            else:
                Rush = "NO"
            #FIN DEL ANALISIS DEL RUSH


            if Total == "0" and (DiferenciaTiempo < 3) and (compararFecha == 1):
                Celda = corrimientoComparacion + 20
                Texto = "COINCIDE ("+ valorAsset +")// Rush:"+Rush
                wSNetflix.cell(FilaNet,Celda, Texto)                                
            else:
                NumeroTareas = "Tareas:" + CompararTarea
                NombreAsset = "Asset: " + str(NombreAsset) + " // "
                Fecha = "Fecha: "+ str(Fecha) + " // "
                Texto = CompararTarea + " // " + NombreAsset + Fecha + "Tiempo:" + str(DiferenciaTiempo) + "min //Rush:"+Rush

                Celda = corrimientoComparacion + 20 
                wSNetflix.cell(FilaNet,Celda, Texto)


            if str(Fecha1) == str(Fecha2):
                Coincide = 1
            corrimientoComparacion =corrimientoComparacion +1
    numeroFila = numeroFila+1   #Incremento del numero de Fila



#################### AUMENTANDO PRECISION #####################

for i in range(15,(UltimaFIla-1)):
    TareaNet = wSNetflix.cell(i,14).value
    TareaNet = TareaNet.strip()
    NumeroOpciones = 0 
    #Buscamos equivalentes de tareas
    TareaOp1 = ''
    TareaOp2 = ''
    filaTarea = 1      

    for row in enumerate(wSTask.rows):               
        if TareaNet == wSTask.cell(filaTarea,1).value:            
            TareaOp1 = wSTask.cell(filaTarea,4).value
            TareaOp2 = wSTask.cell(filaTarea,6).value
            break
        filaTarea = filaTarea + 1  
    #IdAud = wSNetflix.cell(i,14).value
    j = 21

    while wSNetflix.cell(i,j).value != None:
        NumeroOpciones = NumeroOpciones +1
        j = j+1
    if NumeroOpciones > 1:
        for k in range(1,(NumeroOpciones+1)):
            opcion = 20+k
            valorOpcion = wSNetflix.cell(i,opcion).value
            
            ID = wSNetflix.cell(i,opcion).value                         
            ID = ID.strip()
            ID = ID.split('//')
            ID = ID[1]
            ID = ID.split(':')
            ID = ID[1]
            ID = ID.strip()
            NumeroFila = 1
            for row in enumerate(wSBD.rows):
                Celda = wSBD.cell(NumeroFila,1).value 
                Celda = str(Celda).strip()
                if Celda == ID:                           
                    break
                NumeroFila = NumeroFila + 1
            TareasAud = wSBD.cell(NumeroFila,16).value 
            if TareasAud!=None:
                NTareasAud  = TareasAud.split('/')
                for n in NTareasAud:
                    if (n == TareaOp1 or n == TareaOp2):                          
                        wSNetflix.cell(i,21, '')
                        wSNetflix.cell(i,21, valorOpcion)
                        #wSNetflix.cell(i,22, '')
                        break

#Borramos opciones de mas
## CORREGIMOS CUANDO ES VERIFICATION COMPLETE


for i in range(15,(UltimaFIla-1)):
    wSNetflix.cell(i,22, '')
    wSNetflix.cell(i,23, '')
    wSNetflix.cell(i,24, '')
    wSNetflix.cell(i,25, '')
    #Valores
    valores = wSNetflix.cell(i, 21).value
    if valores != None:
        valores = valores.split("//")
        fech1 = wSNetflix.cell(i, 19).value

        if (len(valores) > 2): 
            verificacion = wSNetflix.cell(i, 15).value
            valAsset = valores[1]
            valfech = valores[2]
            valtem = valores[3]
            rush = valores[4]
            fech = valores[2]
            fech = fech.split(":")
            fech = (str(fech[1])).strip()
            fech = fech.split(" ")
            fech = fech[0]
            temp = valores[3]
            temp = temp.split(":")
            temp = (temp[1]).replace("min","")
            temp = int(temp)

            Compara = str(valores[0])
            Compara = Compara.strip()

            index = Compara.find("OK")
            if index != 18:           
                Compara= Compara.split("Tareas")
                Faltantes = Compara[0]
                Compara = Compara[1].replace("[","")
                Compara = Compara.replace("]","")
                Compara = Compara.replace("'","")
                Compara = Compara.replace('"','')
                Compara = Compara.split(",")
                for n in Compara:
                    verificacion = (str(verificacion)).strip()
                    if ((n.strip() == "M&E Full Comparative") and (verificacion=="VERIFICATION_COMPLETE")) or ((n.strip() == "M&E Spot Comparative") and (verificacion=="VERIFICATION_COMPLETE")):                                       
                        Faltantes = Faltantes.strip()
                        if Faltantes == "Faltan 1" or Faltantes == "Faltan -1":
                            AdaptacionAsset = valAsset.replace("Asset:","")
                            AdaptacionAsset = AdaptacionAsset.strip()
                            if (fech == fech1) and (temp>-3 and temp<3):                        
                                Texto = "COINCIDE ("+ AdaptacionAsset +")//" + rush
                                wSNetflix.cell(i,21, Texto)
                            else:
                                Texto = "Numero de tareas: OK //"+valAsset+" // "+valfech + " // "+ valtem +"//"+rush
                                wSNetflix.cell(i,21, Texto)
                            
for i in range(15,(UltimaFIla-1)):    
    valores = wSNetflix.cell(i, 21).value
    if valores != None:
        valores = valores.split("//")    
        if (len(valores) > 2 and valores!=None):         
            Tar = str(valores[0])
            valAsset = valores[1]
            valfech = valores[2]
            valtem = valores[3]
            rush = valores[4]
            Tar = Tar.strip()
            if Tar != "Numero de tareas: OK":            
                Tar = Tar.split("Tareas")
                Tar = (str(Tar[1])).strip()
                Tar = Tar.replace("[","")
                Tar = Tar.replace("]","")
                Tar = Tar.split(",")
                valoresVacios = 0
                for n in Tar:
                    var = n.strip()
                    if var != 'None':
                        valoresVacios = valoresVacios + 1
                if valoresVacios==0:                
                    Texto = "Numero de tareas: OK //"+valAsset+" // "+valfech + " // "+ valtem +"//"+rush
                    wSNetflix.cell(i,21, Texto)
for i in range(15,(UltimaFIla)):
    HojaTarifas  = load_workbook('Tarifas.xlsx')     
    HojaQC = HojaTarifas.worksheets[0] 
    HojaVerificacion = HojaTarifas.worksheets[1]
    minutos = wSNetflix.cell(i,16).value
    valores = wSNetflix.cell(i, 21).value
    if valores != None:
        valores = valores.split("//")
        esRush = valores[(len(valores)-1)]
        esRush = esRush.strip()
        esRush = esRush.replace("Rush:","")
    
        esAudio = wSNetflix.cell(i,14).value
        esAudio = esAudio.find('AUDIO')

        esStream = wSNetflix.cell(i,14).value
        esStream = esStream.find('STREAM')
    
        esFix = wSNetflix.cell(i,14).value
        esFix = esFix.find('FIXCHECK')

        esSpot = wSNetflix.cell(i,14).value
        esSpot = esSpot.find('SPOT')

        esFull = wSNetflix.cell(i,14).value
        esFull = esFull.find('FULL')

        esFlate = wSNetflix.cell(i,17).value
        esFlate = esFlate.strip()
        esFlate = esFlate.split("/")
        esFlate = esFlate[1]

        ratePagado = wSNetflix.cell(i, 17).value
        ratePagado = ratePagado.strip()
        ratePagado = ratePagado.split(" ")
        ratePagado = ratePagado[2]

        tipoTarea = wSNetflix.cell(i, 15).value
        tipoTarea = tipoTarea.strip()

        Tarea = wSNetflix.cell(i, 14).value
        Tarea = Tarea.strip()
        ######## PARA TAREAS TIPO VERIFICACION ###########
        if tipoTarea == "VERIFICATION_COMPLETE":
            filaHVerificacion = 1
            for row in enumerate(HojaVerificacion.rows):
                valOp1 = HojaVerificacion.cell(filaHVerificacion, 2).value
                valOp2 = HojaVerificacion.cell(filaHVerificacion, 3).value           
                Tarifa = HojaVerificacion.cell(filaHVerificacion, 5).value
                if Tarifa!=None:
                    Tarifa = Tarifa.split(" ")
                    Tarifa = Tarifa[0]
                    Tarifa = Tarifa.split(".")
                    Tarifa = Tarifa[0]
                    Tarifa = Tarifa.replace("$","")
                    if (Tarea == valOp1) or (Tarea == valOp2):
                    #if esRush =="SI":
                    #    Tarifa = (int(Tarifa))*1.5
                        Diferencia = float(ratePagado) - float(Tarifa)
                        if (esAudio!=-1)and (esStream!=-1)and(Diferencia==25 or Diferencia==-25)and(esFix!=-1 or esSpot!=-1 or esFull!=-1) :
                            Diferencia=0
                        if Diferencia != 0:
                            Texto = "Se pago: " + str(Diferencia)
                            wSNetflix.cell(i,22, Texto)      
                        break  
                    filaHVerificacion = filaHVerificacion + 1
    ########### PARA TAREAS TIPO QC #################
        else:
            filaQC = 1
            for row in enumerate(HojaQC.rows):        
                valOp1 = HojaQC.cell(filaQC, 2).value
                valOp2 = HojaQC.cell(filaQC, 3).value     
                Tarifa = HojaQC.cell(filaQC, 7).value
                esPorMinuto = HojaQC.cell(filaQC, 7).value
                esPorMinuto = esPorMinuto.strip()
                Tarifa = Tarifa.split(" ")
                Tarifa = Tarifa[0]
                Tarifa = Tarifa.replace("/RT","")
                Tarifa = Tarifa.replace("$","")
                Tarifa = float(Tarifa)
                if (Tarea == valOp1) or (Tarea == valOp2):            
                    minutos = minutos.replace("min","")
                    minutos = minutos.strip()
                    minutos = int(minutos)
                
                    if (esPorMinuto.endswith('minute')) and esFlate=="FLAT_RATE":
                        valorMinuto = Tarifa
                        if minutos <= 15:
                            minutos = 15
                        Tarifa = minutos * valorMinuto


                                
                
                #Por si es Rush
                    if esRush == "SI" and (minutos>15 or esFlate=="FLAT_RATE"):                    
                        Tarifa = Tarifa*1.5                
                    if (esFull!=-1) and (minutos<=15):
                        Tarifa = 108.75
                    Diferencia = (float(ratePagado)-float(Tarifa))

                    if (esAudio!=-1) and (esStream!=-1) and (esSpot!=-1 or esFix!=-1) and(Diferencia==-25 or Diferencia==25):
                        Diferencia = 0
                    if (Diferencia>0.1 )or (Diferencia<-0.1):
                        Texto = "Se pago: " + str(Diferencia)
                        wSNetflix.cell(i,22, Texto)   
                    break
                filaQC = filaQC + 1

def month_string_to_number(string):
    m={
        "Jan":'1',
        "Feb":'2',
        "Mar":'3',
        "Apr":'4',
        "May":'5',
        "Jun":'6',
        "Jul":'7',
        "Aug":'8',
        "Sep":'9',
        "Oct":'10',
        "Nov":'11',
        "Dec":'12'
    }
    s = string.strip()
    s = s.capitalize()
    try: 
        out = m[s]
        return out
    except:
        raise ValueError("Mes no Encontrados")

FilAudiomaster = 2
FechaEn = FechaMenor
FechaFin = FechaMayor
FechaEn = datetime.strptime(FechaEn,'%Y-%m-%d')
FechaFin = datetime.strptime(FechaFin,'%Y-%m-%d')

FilaNetflix = UltimaFIla + 20
Faltantes = {'Faltantes'}
Existentes = {'Existen'}

while wSBD.cell(FilAudiomaster,1).value != None:
    Fecha = wSBD.cell(FilAudiomaster,19).value
    UID  = wSBD.cell(FilAudiomaster,25).value
    re =''
    Net =''
    if UID !=None:
        re = UID.find('requestID');
        Net = UID.find('LOLI'); 
    else: 
        re=-1
        Net = -1
 
    if re != -1 or Net != -1:
        if Net != -1:
            UID = UID.split("LOLI")
        if  re != -1:
            UID = UID.split("requestID")
        UID = UID[1]        
        UID = UID.strip()
        AssetAudiomaster = wSBD.cell(FilAudiomaster,1).value
        TituloAudiomaster = wSBD.cell(FilAudiomaster,3).value
        MovieIdAudiomaster = wSBD.cell(FilAudiomaster,5).value
        showID = "None"
        titleType = "None"
        FileLocationAudiomaster = wSBD.cell(FilAudiomaster,25).value
        PackageIdAudiomaster = wSBD.cell(FilAudiomaster,24).value
        OperadorAudiomaster = wSBD.cell(FilAudiomaster,26).value
        OpAccuentAudiomaster = wSBD.cell(FilAudiomaster,23).value
        TareaAudiomaster = wSBD.cell(FilAudiomaster,16).value
        QCOPerador = wSBD.cell(FilAudiomaster,6).value
        DuracionAudiomaster = wSBD.cell(FilAudiomaster,7).value
        RushAudiomaster = wSBD.cell(FilAudiomaster,15).value
        LanguageAudiomaster = wSBD.cell(FilAudiomaster,27).value
        FechaAudiomaster = wSBD.cell(FilAudiomaster,28).value
        Estatus = wSBD.cell(FilAudiomaster,18).value
     
        Fecha = str(Fecha).replace("00:00:00","")
        print("->", FilAudiomaster, "--->", Fecha, "-->", FechaEn,"/", FechaFin)
        if str(Fecha).find('-') != (-1):
            Fecha = Fecha.strip()
            Fecha = Fecha.split("-")
            Fecha = Fecha[0]+"-"+Fecha[1]+"-"+Fecha[2]            
            Fecha = Fecha.strip()
            Fecha = datetime.strptime(Fecha,'%Y-%m-%d')
            
            if (Fecha >= FechaEn ) and (Fecha <= FechaFin):      
                FilaAna = 15                            
                while wSNetflix.cell(FilaAna,3).value != None:                    
                    valCeld = wSNetflix.cell(FilaAna,3).value
                    valCeld = valCeld.split("||")
                    if len(valCeld) == 2:
                        valCeld = valCeld[0]
                        valCeld = valCeld.split(" ")
                        valCeld = valCeld[1]
                        valCeld = valCeld.replace("(","")
                        valCeld = valCeld.replace(")","")
                        valCeld = valCeld.strip()                    
                    else:
                        valCeld = valCeld[4]
                        valCeld = valCeld.replace("Request ID -","")
                        valCeld = str(valCeld).strip()     
                    UID = UID.replace("=","")
                    UID = UID.strip()
                    if (UID == valCeld):
                        print("---->",FilAudiomaster,"--",UID,"<---2-->", valCeld)
                        Tex = str(AssetAudiomaster)+"||"+str(TituloAudiomaster)+"||"+str(MovieIdAudiomaster)+"||"+showID+"||"+titleType+"||"+str(FileLocationAudiomaster)+ "||"+ str(PackageIdAudiomaster) +"||"+str(OperadorAudiomaster)+"||"+str(OpAccuentAudiomaster)+"||"+str(TareaAudiomaster)+"||"+str(QCOPerador)+"||"+str(DuracionAudiomaster)+"||"+str(RushAudiomaster)+"||"+str(LanguageAudiomaster)+"||"+str(FechaAudiomaster)+"||"+Estatus+"||"+str(RushAudiomaster)
                        Existentes.add(Tex)                                        
                        break
                    else:
                        Tex = str(AssetAudiomaster)+"||"+str(TituloAudiomaster)+"||"+str(MovieIdAudiomaster)+"||"+showID+"||"+titleType+"||"+str(FileLocationAudiomaster)+ "||"+ str(PackageIdAudiomaster) +"||"+str(OperadorAudiomaster)+"||"+str(OpAccuentAudiomaster)+"||"+str(TareaAudiomaster)+"||"+str(QCOPerador)+"||"+str(DuracionAudiomaster)+"||"+str(RushAudiomaster)+"||"+str(LanguageAudiomaster)+"||"+str(FechaAudiomaster)+"||"+Estatus+"||"+str(RushAudiomaster)
                        Faltantes.add(Tex)                   
                        FilaNetflix = FilaNetflix+1
                    FilaAna += 1
    FilAudiomaster = FilAudiomaster +1
filFaltantes = 1
wSNetflix.cell(UltimaFIla+15,1, "Proyectos Faltantes:")   
for n in (Faltantes-Existentes):
    if n != "Faltantes":
        wSNetflix.cell(UltimaFIla+15+filFaltantes,1, n)  
        filFaltantes += 1
filaDesglose = UltimaFIla + 16
while wSNetflix.cell(filaDesglose,1).value != None:
    HojaTarifas  = load_workbook('Tarifas.xlsx')     
    HojaQC = HojaTarifas.worksheets[0] 
    HojaVerificacion = HojaTarifas.worksheets[1]

    Desglozar = wSNetflix.cell(filaDesglose,1).value
    Desglozar = Desglozar.split("||")
    ColumnaDesglose = 2
    for n in Desglozar:
        wSNetflix.cell(filaDesglose,ColumnaDesglose, n) 

        ColumnaDesglose += 1
        
    TareasFaltantes = wSNetflix.cell(filaDesglose,9).value
    TareasFaltantes = TareasFaltantes.split("/")
    Tipo = wSNetflix.cell(filaDesglose,10).value
    for t in TareasFaltantes:
        filTarifa = 1
        Tipo = (str(Tipo)).strip()
        
        if Tipo=="QCer":
            for row in enumerate(HojaQC.rows):
                t.strip()
                tarea2 = HojaQC.cell(filTarifa,4).value
                tarea2.strip()
                
                if t==tarea2:
                    CostoN = HojaQC.cell(filTarifa,7).value
                    Costo = wSNetflix.cell(filTarifa,20).value
                    Costo = str(Costo)+"+"+str(CostoN)
                    wSNetflix.cell(filaDesglose,20, Costo)
                    break
                filTarifa += 1
        if Tipo=="Verifier":
            for row in enumerate(HojaVerificacion.rows):
                tarea2 = HojaVerificacion.cell(filTarifa,4).value
                t.strip()
                
                if t==tarea2 and tarea2!=None:
                    tarea2.strip()
                    CostoN = HojaQC.cell(filTarifa,5).value
                    Costo = wSNetflix.cell(filTarifa,20).value
                    Costo = str(Costo)+"+"+str(CostoN)
                    wSNetflix.cell(filaDesglose,20, Costo)
                    break
                filTarifa += 1
    filaDesglose += 1
################ FIN DE AUMENTO DE PRECISION ##############

print("Fecha Mayor: ", FechaMayor,"---  Fecga Menor: ", FechaMenor);
    ################ FIN DE ANALISIS DE ARCHIVO     ############################################    
# GUARDAMOS EL ARCHIVO FINAL CON EL NOMBRE DE ARCHIVO
wbNetflix.save(filename="Analisis.xlsx")

